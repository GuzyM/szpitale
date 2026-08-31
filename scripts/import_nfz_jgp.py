#!/usr/bin/env python3
"""Convert official NFZ attachments 1 and 9 into HospitalAPP JavaScript shards."""
import argparse, json, math, re
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parents[1]
SOURCE = "https://baw.nfz.gov.pl/NFZ/document/43942/Zarzadzenie-74_2026_DSOZ"
MEDICAL = re.compile(r"^[A-Z0-9][A-Z0-9.\-+*]{0,12}\s+\S", re.I)
REF = re.compile(r"list(?:y|a|ę|ą)\s+(?:procedur|rozpoznań|dodatkowej|ogólnej)?\s*([A-ZŁŚŻŹĆŃ0-9]{1,12})", re.I)

def clean(v):
    return " ".join(v.split()) or None if isinstance(v, str) else v

def num(v):
    v = clean(v)
    if v in (None, ""): return None
    try:
        n = float(str(v).replace(",", ".")); return int(n) if n.is_integer() else n
    except ValueError: return None

def read_catalog(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    groups, scope_labels = [], set()
    for ws in wb.worksheets:
        short = "1ae" in ws.title.lower(); start = 10 if short else 12; note = 78 if short else 80
        labels = [clean(c.value) for c in ws[4]]; section = None
        for row in ws.iter_rows(min_row=6, values_only=True):
            order, code = num(row[0]), clean(row[1])
            if not order or not code:
                if clean(row[0]): section = clean(row[0])
                continue
            scopes=[]
            for i in range(start, note):
                q = num(row[i]) if i < len(row) else None; label = labels[i] if i < len(labels) else None
                if q is not None and label:
                    scopes.append({"label":label,"qualifier":q}); scope_labels.add(label)
            groups.append({
                "order":len(groups)+1,"code":code,"productCode":clean(row[2]),"name":clean(row[3]),"section":section,
                "ordinary":num(row[4]),"planned":num(row[5]),"oneDayTreatment":num(row[6]),
                "under12h":num(row[7]) if short else None,"financedDays":num(row[8] if short else row[7]),
                "sameDay":None if short else num(row[8]),"oneDayHosp":None if short else num(row[9]),
                "twoDayHosp":None if short else num(row[10]),"extraDay":num(row[9] if short else row[11]),
                "scopeFamilies":scopes,"catalogNote":clean(row[note]) if note < len(row) else None,
                "catalogVariant":"1ae" if short else "1a"})
    return groups, len(scope_labels)

def role(text):
    text=text.lower()
    return "procedure" if "procedur" in text else "diagnosis" if "rozpozna" in text else "general" if "ogóln" in text else "additional" if "dodatkow" in text else "reference"

def read_blocks(path, group_codes):
    wb=openpyxl.load_workbook(path,read_only=True,data_only=True); blocks={}; sections={}
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=4,values_only=True):
            section,code,value=clean(row[0]),clean(row[1]),clean(row[2])
            if not value: continue
            if not code:
                if section and str(value).startswith(section+" "): sections[section]=str(value)
                continue
            code=str(code); value=str(value)
            b=blocks.setdefault(code,{"section":section,"kind":"group" if code in group_codes else "list","title":code,"segments":[],"references":[]})
            if value==code or value.startswith(code+" "):
                if b["title"]==code: b["title"]=value
                continue
            if value in ("ICD-9","ICD-10"):
                b["segments"].append({"type":"list","label":value,"system":value,"items":[]}); continue
            current=b["segments"][-1] if b["segments"] else None
            if current and current["type"]=="list" and MEDICAL.match(value): current["items"].append(value); continue
            b["segments"].append({"type":"text","text":value})
            seen={(x["code"],x["role"]) for x in b["references"]}
            for m in REF.finditer(value):
                rc=m.group(1).upper(); rr=role(m.group(0))
                if rc!=code.upper() and (rc,rr) not in seen: b["references"].append({"code":rc,"role":rr}); seen.add((rc,rr))
    for b in blocks.values(): b["segments"]=[s for s in b["segments"] if s["type"]!="list" or s["items"]]
    return blocks,sections

def shard(items,count):
    size=max(1,math.ceil(len(items)/count)); return [items[i:i+size] for i in range(0,len(items),size)]

def js(v): return json.dumps(v,ensure_ascii=False,separators=(",",":"))

def main():
    ap=argparse.ArgumentParser(); ap.add_argument("catalog",type=Path); ap.add_argument("characteristics",type=Path); a=ap.parse_args()
    groups,scopes=read_catalog(a.catalog); codes={g["code"] for g in groups}; blocks,sections=read_blocks(a.characteristics,codes); out=ROOT/"data"
    meta={"meta":{"orderNumber":"74/2026/DSOZ","orderDate":"2026-07-23","effectiveFrom":"2026-07-01","attachment":"Załącznik 1 — katalogi grup 1a i 1ae","catalog":"1a oraz 1ae — pobyt do 12 godzin","sourceUrl":SOURCE,"groupCount":len(groups),"scopeFamilyCount":scopes,"status":"official"},"groups":[]}
    (out/"jgp-data-meta.js").write_text(f"window.JGP_CATALOG = {js(meta)};\n",encoding="utf-8")
    chunks=shard(groups,4)
    for i in range(4): (out/f"jgp-data-{i+1:02}.js").write_text(f"window.JGP_CATALOG.groups.push(...{js(chunks[i] if i<len(chunks) else [])});\n",encoding="utf-8")
    entries=sum(len(s["items"]) for b in blocks.values() for s in b["segments"] if s["type"]=="list")
    cm={"meta":{"orderNumber":"74/2026/DSOZ","orderDate":"2026-07-23","effectiveFrom":"2026-07-01","attachment":"Załącznik 9 — Charakterystyka JGP 1a i 1ae","sourceUrl":SOURCE,"groupCount":len(codes),"listCount":sum(b["kind"]=="list" for b in blocks.values()),"codeEntryCount":entries,"status":"official"},"blocks":{},"sectionTitles":sections}
    (out/"jgp-characteristics-meta.js").write_text(f"window.JGP_CHARACTERISTICS = {js(cm)};\n",encoding="utf-8")
    chunks=shard(sorted(blocks.items()),14)
    for i in range(14): (out/f"jgp-characteristics-{i+1:02}.js").write_text(f"Object.assign(window.JGP_CHARACTERISTICS.blocks,{js(dict(chunks[i]) if i<len(chunks) else {})});\n",encoding="utf-8")
    print({"groups":len(groups),"blocks":len(blocks),"codes":entries})

if __name__=="__main__": main()
